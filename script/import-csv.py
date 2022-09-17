#!/usr/bin/env python

import csv
import os
import sys
import urllib.request

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


os.chdir(sys.path[0])
if not os.path.exists("output"):
    os.mkdir("output")

sheet_id = "1Msr7Vdqfa6MlMXYSlERAnbTwEr4Fd7rVe2uvxaPjo14"
sheet_url = (
    f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx&gid=0"
)

xlsx_file = "memento.xlsx"
urllib.request.urlretrieve(sheet_url, xlsx_file)

wb = load_workbook(filename=xlsx_file)
ws = wb.active

while ws.merged_cells:
    for merged in ws.merged_cells.ranges:
        ws.unmerge_cells(str(merged))
        min_col, min_row, max_col, max_row = range_boundaries(str(merged))

        for row in ws.iter_rows(
            min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
        ):
            for cell in row:
                cell.value = merged.start_cell.value

with open("output/input.csv", "w", newline="") as fp:
    csv_writer = csv.writer(fp, lineterminator="\n")
    col_values = [
        first_col for (first_col,) in ws.iter_rows(max_col=1, values_only=True)
    ]
    start_idx = 0

    try:
        start_idx = col_values.index("Season")
    except ValueError as ex:
        print(ex)

    for row in ws.iter_rows(min_row=start_idx + 1, values_only=True):
        if all([cell == None for cell in row]):
            continue

        csv_writer.writerow(row)

os.remove(xlsx_file)
